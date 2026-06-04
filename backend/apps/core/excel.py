from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="EAF3F8")
TITLE_FILL = PatternFill("solid", fgColor="007CB7")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9E3"),
    right=Side(style="thin", color="D9D9E3"),
    top=Side(style="thin", color="D9D9E3"),
    bottom=Side(style="thin", color="D9D9E3"),
)
TRUE_VALUES = {"1", "true", "t", "yes", "y", "si", "s", "con", "items", "with"}


def query_bool(request, name="with_items"):
    return str(request.query_params.get(name, "")).strip().lower() in TRUE_VALUES


def _filename(value):
    return "".join(char if char.isalnum() or char in {"-", "_"} else "_" for char in str(value))


def _sheet_title(value):
    cleaned = "".join(char for char in str(value) if char not in r"[]:*?/\\").strip() or "Datos"
    return cleaned[:31]


def _normalize(value):
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return float(value)
    if hasattr(value, "tzinfo") and value.tzinfo:
        return timezone.localtime(value).replace(tzinfo=None)
    return value


def build_workbook(title, headers, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _sheet_title(title)

    sheet.append([title])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
    sheet["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    sheet["A1"].fill = TITLE_FILL
    sheet["A1"].alignment = Alignment(horizontal="center")

    generated_at = timezone.localtime(timezone.now()).replace(tzinfo=None)
    sheet.append(["Generado", generated_at])
    sheet.append([])
    sheet.append(list(headers))
    header_row = 4

    for row in rows:
        sheet.append([_normalize(value) for value in row])

    for cell in sheet[header_row]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    for row in sheet.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top")
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

    for column_index in range(1, len(headers) + 1):
        letter = get_column_letter(column_index)
        max_length = 12
        for cell in sheet[letter]:
            max_length = max(max_length, len(str(cell.value or "")))
        sheet.column_dimensions[letter].width = min(max_length + 2, 42)

    sheet.freeze_panes = "A5"
    if rows:
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{sheet.max_row}"
    return workbook


def excel_response(workbook, filename):
    buffer = BytesIO()
    workbook.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{_filename(filename)}.xlsx"'
    return response
